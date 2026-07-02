import io
from pathlib import Path

import pytest
from openpyxl import Workbook
from sentier_importers.core import parse as parse_mod
from sentier_importers.core.errors import ParseError
from sentier_importers.core.types import RawData

FIXTURES = Path(__file__).parent / "fixtures"


def _raw(text: str) -> RawData:
    return RawData(content=text.encode("utf-8"), source_url="file:///x")


def _xlsx_bytes(workbook: Workbook) -> RawData:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return RawData(content=buffer.getvalue(), source_url="file:///x.xlsx")


def test_parse_csv():
    records = parse_mod.parse(_raw("id,name\n1,alpha\n2,beta\n"), "csv")
    assert records == [{"id": "1", "name": "alpha"}, {"id": "2", "name": "beta"}]


def test_parse_json_array():
    records = parse_mod.parse(_raw('[{"a": 1}, {"a": 2}]'), "json")
    assert records == [{"a": 1}, {"a": 2}]


def test_parse_json_object_wrapped():
    records = parse_mod.parse(_raw('{"a": 1}'), "json")
    assert records == [{"a": 1}]


def test_parse_yaml():
    records = parse_mod.parse(_raw("- a: 1\n- a: 2\n"), "yaml")
    assert records == [{"a": 1}, {"a": 2}]


def test_parse_ttl():
    ttl = "@prefix ex: <http://example.org/> .\nex:s ex:p ex:o .\n"
    records = parse_mod.parse(_raw(ttl), "ttl")
    assert records == [
        {
            "subject": "http://example.org/s",
            "predicate": "http://example.org/p",
            "object": "http://example.org/o",
        }
    ]


def test_unknown_format_raises():
    with pytest.raises(ParseError):
        parse_mod.parse(_raw("x"), "toml")


def test_register_parser_extends_registry():
    parse_mod.register_parser("upper", lambda raw: [{"v": raw.content.decode().upper()}])
    assert parse_mod.parse(_raw("hi"), "upper") == [{"v": "HI"}]


def test_parse_xlsx_fixture_roundtrip():
    raw = RawData(content=(FIXTURES / "sample.xlsx").read_bytes(), source_url="file:///x.xlsx")
    records = parse_mod.parse(raw, "xlsx")
    # Excel has no pure date type; a date cell round-trips as midnight datetime.
    assert records == [
        {"code": "A1", "label": "Alpha", "updated": "2026-06-05T12:30:00"},
        {"code": "B2", "label": "Beta", "updated": "2026-01-01T00:00:00"},
    ]


def test_parse_xlsx_alias_xls():
    raw = RawData(content=(FIXTURES / "sample.xlsx").read_bytes(), source_url="file:///x.xls")
    assert len(parse_mod.parse(raw, "xls")) == 2


def test_parse_xlsx_skips_empty_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([None, None])  # skipped
    ws.append([1, 2])
    records = parse_mod.parse(_xlsx_bytes(wb), "xlsx")
    assert records == [{"a": 1, "b": 2}]


def test_parse_xlsx_omits_none_cells():
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, None])  # b omitted, not empty string
    records = parse_mod.parse(_xlsx_bytes(wb), "xlsx")
    assert records == [{"a": 1}]


def test_parse_xlsx_trims_header_whitespace():
    wb = Workbook()
    ws = wb.active
    ws.append(["  code ", " label"])
    ws.append(["x", "y"])
    records = parse_mod.parse(_xlsx_bytes(wb), "xlsx")
    assert records == [{"code": "x", "label": "y"}]


def test_parse_xlsx_skips_leading_blank_rows_to_find_header():
    wb = Workbook()
    ws = wb.active
    ws.append([None, None])
    ws.append(["a", "b"])
    ws.append([1, 2])
    records = parse_mod.parse(_xlsx_bytes(wb), "xlsx")
    assert records == [{"a": 1, "b": 2}]


def test_parse_xlsx_empty_workbook_raises():
    wb = Workbook()  # active sheet has no rows
    with pytest.raises(ParseError):
        parse_mod.parse(_xlsx_bytes(wb), "xlsx")


def test_parse_xlsx_duplicate_headers_raise():
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "a"])
    ws.append([1, 2])
    with pytest.raises(ParseError):
        parse_mod.parse(_xlsx_bytes(wb), "xlsx")
